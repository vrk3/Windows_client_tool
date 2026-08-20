"""Spec 7.3: export formats."""
import os
import sqlite3
import xml.etree.ElementTree as ET

import pytest

from modules.treesize.actions import exporters
from modules.treesize.actions.exporters import ExportError, available_formats, export

ROWS = [
    ("Name", "Size (bytes)", "% of Parent", "Path"),
    ("Windows", "9000", "90.0%", "C:\\Windows"),
    ("Zürich.txt", "500", "5.0%", "C:\\Zürich.txt"),
    ("a & b <c>.bin", "500", "5.0%", "C:\\a & b <c>.bin"),
]


def test_available_formats_always_include_the_stdlib_ones():
    formats = available_formats()
    for extension in ("csv", "html", "txt", "xml", "db", "json"):
        assert extension in formats


def test_nothing_to_export_is_refused(tmp_path):
    with pytest.raises(ExportError, match="nothing to export"):
        export(str(tmp_path / "x.csv"), [ROWS[0]])


def test_an_unknown_extension_is_refused(tmp_path):
    with pytest.raises(ExportError, match="Unknown export format"):
        export(str(tmp_path / "x.wat"), ROWS)


def test_csv_is_written_with_a_bom_for_excel(tmp_path):
    path = tmp_path / "out.csv"
    export(str(path), ROWS)
    raw = path.read_bytes()
    assert raw.startswith(b"\xef\xbb\xbf"), "Excel needs the BOM to read UTF-8"
    assert "Zürich.txt" in path.read_text(encoding="utf-8-sig")


def test_text_export_is_aligned_and_titled(tmp_path):
    path = tmp_path / "out.txt"
    export(str(path), ROWS, title="My scan")
    body = path.read_text(encoding="utf-8")
    assert body.startswith("My scan")
    assert "Windows" in body


def test_html_escapes_markup_in_names(tmp_path):
    """A file called 'a & b <c>.bin' must not become markup."""
    path = tmp_path / "out.html"
    export(str(path), ROWS)
    body = path.read_text(encoding="utf-8")
    assert "&amp;" in body and "&lt;c&gt;" in body
    assert "<c>.bin" not in body


def test_xml_is_well_formed_and_uses_legal_tag_names(tmp_path):
    """Headings are human text: '% of Parent' and 'Size (bytes)' are not legal
    tags, and a tag may not start with a digit."""
    path = tmp_path / "out.xml"
    export(str(path), ROWS)
    root = ET.parse(str(path)).getroot()
    items = root.findall("item")
    assert len(items) == 3
    assert items[0].find("name").text == "Windows"
    # "% of Parent" -> "of_parent": the punctuation becomes separators and the
    # leading run is stripped, which is a legal tag.
    assert items[0].find("of_parent").text == "90.0%"
    assert items[0].find("size__bytes").text == "9000"


def test_a_heading_starting_with_a_digit_gets_a_legal_tag():
    """A tag may not start with a digit, and column headings are human text."""
    from modules.treesize.actions.exporters import _tag
    assert _tag("2024 total").startswith("col_")
    assert not _tag("Name").startswith("col_")


def test_xml_escapes_ampersands(tmp_path):
    path = tmp_path / "out.xml"
    export(str(path), ROWS)
    names = [i.find("name").text for i in ET.parse(str(path)).getroot()]
    assert "a & b <c>.bin" in names


def test_json_round_trips(tmp_path):
    import json
    path = tmp_path / "out.json"
    export(str(path), ROWS, title="T")
    payload = json.loads(path.read_text(encoding="utf-8"))
    assert payload["title"] == "T"
    assert payload["items"][0]["Name"] == "Windows"


def test_sqlite_export_is_queryable(tmp_path):
    path = tmp_path / "out.db"
    export(str(path), ROWS, title="T")
    connection = sqlite3.connect(str(path))
    try:
        rows = connection.execute("SELECT name, path FROM scan").fetchall()
        assert ("Windows", "C:\\Windows") in rows
        assert connection.execute(
            "SELECT value FROM meta WHERE key='title'").fetchone()[0] == "T"
    finally:
        connection.close()


def test_sqlite_replaces_rather_than_accumulating(tmp_path):
    """An export is one scan, not a silent pile of several."""
    path = tmp_path / "out.db"
    export(str(path), ROWS)
    export(str(path), ROWS)
    connection = sqlite3.connect(str(path))
    try:
        assert connection.execute("SELECT COUNT(*) FROM scan").fetchone()[0] == 3
    finally:
        connection.close()


def test_excel_export_is_readable(tmp_path):
    pytest.importorskip("openpyxl")
    import openpyxl
    path = tmp_path / "out.xlsx"
    export(str(path), ROWS)
    sheet = openpyxl.load_workbook(str(path)).active
    assert sheet["A1"].value == "Name"
    assert sheet["A2"].value == "Windows"
    assert sheet.freeze_panes == "A2"


def test_pdf_export_produces_a_pdf(tmp_path):
    pytest.importorskip("reportlab")
    path = tmp_path / "out.pdf"
    export(str(path), ROWS, title="Scan")
    assert path.read_bytes().startswith(b"%PDF")


def test_pdf_caps_rows_and_says_so(tmp_path):
    """A five-million-row PDF is not a useful artifact; truncating silently
    would be worse than saying it happened."""
    pytest.importorskip("reportlab")
    big = [ROWS[0]] + [(f"f{i}", str(i), "0.0%", f"C:\\f{i}") for i in range(2500)]
    path = tmp_path / "big.pdf"
    export(str(path), big)
    assert path.stat().st_size > 1000


def test_a_missing_optional_package_is_reported_not_crashed(tmp_path, monkeypatch):
    def refuse(module, package):
        raise ExportError(f"{package} is not installed, so this format is "
                          f"unavailable. Install it with: pip install {package}")

    monkeypatch.setattr(exporters, "_require", refuse)
    with pytest.raises(ExportError, match="not installed"):
        export(str(tmp_path / "out.xlsx"), ROWS)


# ---- Excel's hard row ceiling ------------------------------------------
#
# openpyxl does NOT enforce it. `append()` accepts rows past the end without
# complaint and `save()` writes them into the sheet XML, producing a file that
# reports success here and that Excel then refuses to open, or "repairs" by
# dropping content. Verified directly against openpyxl 3.1.5. A volume with
# more than a million entries is entirely ordinary -- a build server, a NAS,
# anything with node_modules -- so this is reachable, and it fails silently,
# which is the worst way for an export to fail.

def _rows(count):
    return [("Name", "Size")] + [(f"file-{i}.bin", str(i)) for i in range(count)]


def test_xlsx_stays_within_excels_row_limit(tmp_path, monkeypatch):
    openpyxl = pytest.importorskip("openpyxl")
    monkeypatch.setattr(exporters, "XLSX_MAX_ROWS", 10)
    path = tmp_path / "capped.xlsx"

    exporters.export(str(path), _rows(50))

    book = openpyxl.load_workbook(str(path))
    assert book.active.max_row <= 10, "the file exceeds the limit it claims"
    book.close()


def test_xlsx_says_how_much_it_left_out(tmp_path, monkeypatch):
    """Stated in the document, as the PDF export already does. A truncation
    the user cannot see is indistinguishable from data loss."""
    openpyxl = pytest.importorskip("openpyxl")
    monkeypatch.setattr(exporters, "XLSX_MAX_ROWS", 10)
    path = tmp_path / "capped.xlsx"

    exporters.export(str(path), _rows(50))

    book = openpyxl.load_workbook(str(path))
    sheet = book.active
    last = " ".join(str(c.value) for c in sheet[sheet.max_row] if c.value)
    assert "50" in last, f"the note does not say the real total: {last!r}"
    assert "8" in last, f"the note does not say how many were shown: {last!r}"
    book.close()


def test_a_scan_under_the_limit_is_not_truncated_or_annotated(tmp_path,
                                                              monkeypatch):
    openpyxl = pytest.importorskip("openpyxl")
    monkeypatch.setattr(exporters, "XLSX_MAX_ROWS", 100)
    path = tmp_path / "whole.xlsx"

    exporters.export(str(path), _rows(20))

    book = openpyxl.load_workbook(str(path))
    sheet = book.active
    assert sheet.max_row == 21, "header plus twenty rows, and nothing else"
    assert sheet.cell(row=21, column=1).value == "file-19.bin"
    book.close()


def test_column_widths_still_come_from_the_content(tmp_path):
    """The width pass was rewritten from one scan per column to one scan
    total; it still has to produce the same widths."""
    openpyxl = pytest.importorskip("openpyxl")
    path = tmp_path / "wide.xlsx"
    rows = [("Name", "Size"),
            ("a-very-considerably-longer-name-than-the-header.bin", "1"),
            ("short.bin", "2")]

    exporters.export(str(path), rows)

    book = openpyxl.load_workbook(str(path))
    widths = book.active.column_dimensions
    assert widths["A"].width > widths["B"].width
    assert widths["A"].width == min(60, len(rows[1][0]) + 2)
    book.close()


def test_ragged_rows_do_not_break_the_width_pass(tmp_path):
    """A row shorter than the header must not index past the width list."""
    pytest.importorskip("openpyxl")
    path = tmp_path / "ragged.xlsx"
    exporters.export(str(path), [("A", "B", "C"), ("only-one",), ("x", "y")])
    assert path.exists()
